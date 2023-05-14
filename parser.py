#!/usr/bin/env python

import argparse
import datetime
import pathlib
import re
import sqlite3
from typing import Generator

import openpyxl


def as_int(value) -> int:
    try:
        return int(re.sub(r'.*?(\d+)$', r'\1', str(value)))
    except (TypeError, ValueError):
        return 0


class ReportTable:

    __fields = dict(
        type='TEXT',
        name='TEXT',
        date='TEXT',
        value='INTEGER',
        company_id='INTEGER',
    )

    def __init__(self, filename: str, table: str = 'report'):
        self.conn, self.table = sqlite3.connect(filename), table
        self._create()

    def _create(self) -> None:
        self._execute(f'DROP TABLE IF EXISTS {self.table}')
        fields = tuple(
            f'{name} {type} NOT NULL' for name, type in self.__fields.items()
        )
        self._execute(
            f'CREATE TABLE IF NOT EXISTS {self.table} ({", ".join(fields)})',
        )

    def _execute(self, sql: str) -> None:
        self.conn.execute(sql)
        self.conn.commit()

    def append(
        self,
        type: str, name: str, date: datetime.date, value: int, company_id: int
    ) -> None:
        self._execute(f"""INSERT INTO {self.table} VALUES (
            '{type}',
            '{name}',
            '{date.strftime("%Y-%m-%d")}',
            {value},
            {company_id}
        )""")

    def select(self, *fields, group_by: str | None = None)\
            -> Generator[dict[str, str | int], None, None]:
        sql = f'SELECT {", ".join(map(str, fields))} FROM {self.table}'
        if group_by:
            sql += f' GROUP BY {group_by}'
        cursor = self.conn.execute(sql)
        names = tuple(desc[0] for desc in cursor.description)
        for row in cursor:
            yield dict(zip(names, row))


class SheetParser:

    DataTuple = tuple[
        str,  # indicator type
        str,  # indicator name
        datetime.date,  # value date
        int,  # indicator value
        int,  # company identifier
    ]

    def __init__(self, filename: str, sheet_name: str | None = None):
        workbook = openpyxl.load_workbook(filename)
        self.sheet = workbook[sheet_name] if sheet_name else workbook.active
        self.max_column = None

        for i in range(0, self.sheet.max_column):
            cell = self.sheet.cell(1, i + 1)
            if cell.value is None and not self.is_merged(cell):
                break
            self.max_column = re.sub(r'^(\w+)\d+$', r'\1', cell.coordinate)

    @staticmethod
    def __coord(row: int, col: str) -> str:
        return col + str(row)

    def get_cell_value(self, row: int, col: int, default: str | None)\
            -> str | None:
        cell = self.sheet.cell(row, col)

        return default if cell.value is None and self.is_merged(cell) \
            else str(cell.value).lower()

    def is_merged(self, cell) -> bool:
        """
        :type cell: openpyxl.Cell
        """
        for merged_range in self.sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False

    def iter(
        self, start_row: int = 4, start_col: str = 'C',
        type_row: int = 1, name_row: int = 2, date_row: int = 3,
        company_col: str = 'B'
    ) -> Generator[DataTuple, None, None]:
        start = self.__coord(start_row, start_col)
        end = self.__coord(self.sheet.max_row, self.max_column)
        type = name = date = None
        today = datetime.date.today()

        for row in self.sheet[start:end]:
            for cell in row:
                type = self.get_cell_value(type_row, cell.column, type)
                name = self.get_cell_value(name_row, cell.column, name)
                date = self.get_cell_value(date_row, cell.column, date)
                ccell = self.sheet[self.__coord(cell.row, company_col)]
                yield (
                    type,
                    name,
                    today + datetime.timedelta(days=as_int(date)),
                    as_int(cell.value),
                    as_int(ccell.value),
                )


arg_parser = argparse.ArgumentParser()
arg_parser.add_argument('input', help='Input Excel file name')
arg_parser.add_argument('--output', '-o', help='Output SQLite file name')
args = arg_parser.parse_args()

if not (report := args.output):
    path = pathlib.Path(args.input)
    report = str(path.parent / f'{path.stem}.db')
report = ReportTable(report)
for row in SheetParser(args.input).iter():
    report.append(*row)
report = report.select(
    'date', 'name indicator', 'type', 'SUM(value) total',
    group_by='date, name, type',
)

columns = 'date', 'indicator', 'type', 'total'
print('\t'.join('{:10}'.format(col) for col in columns))
for row in report:
    print('\t'.join('{:10}'.format(row[col]) for col in columns))
